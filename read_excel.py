"""
اسکریپت برای خواندن و تحلیل فایل Excel
"""
import openpyxl
from openpyxl.utils import get_column_letter

# باز کردن فایل با فرمول‌ها
wb = openpyxl.load_workbook(r'c:\Users\Moein\Desktop\Raw invoice.xlsm', keep_vba=True, data_only=False)

print("=" * 80)
print("تحلیل فایل Raw invoice.xlsm")
print("=" * 80)

# بررسی هر دو Sheet
for sheet_name in wb.sheetnames:
    print(f"\n{'=' * 60}")
    print(f"### Sheet: {sheet_name} ###")
    print('=' * 60)
    ws = wb[sheet_name]
    
    # نمایش هدرها (سطر 6)
    print("\nهدرهای سطر 6:")
    headers = {}
    for col in range(1, 36):
        cell = ws.cell(6, col)
        if cell.value:
            col_letter = get_column_letter(col)
            headers[col_letter] = cell.value
            print(f"  {col_letter}: {cell.value}")
    
    # بررسی چند سطر برای یافتن فرمول‌ها و داده‌ها
    print("\nبررسی سطرهای 7 تا 20 برای فرمول‌ها:")
    for row in range(7, 21):
        row_has_content = False
        row_data = []
        
        for col in range(1, 36):
            cell = ws.cell(row, col)
            col_letter = get_column_letter(col)
            
            # بررسی فرمول
            if cell.data_type == 'f':
                if not row_has_content:
                    print(f"\n  سطر {row}:")
                    row_has_content = True
                print(f"    {col_letter}{row} = {cell.value}")
                if cell.value and isinstance(cell.value, str):
                    print(f"      Formula: {cell.value}")
            elif cell.value and cell.value != '' and not isinstance(cell.value, (int, float)) or (isinstance(cell.value, (int, float)) and cell.value != 0):
                row_data.append(f"{col_letter}={cell.value}")
        
        if row_data and len(row_data) > 2:
            print(f"\n  سطر {row} داده: {' | '.join(row_data[:10])}")
    
    # بررسی یک سطر نمونه با جزئیات کامل
    print(f"\n\nجزئیات کامل سطر 7:")
    for col_letter, header in headers.items():
        col_num = openpyxl.utils.column_index_from_string(col_letter)
        cell = ws.cell(7, col_num)
        print(f"  {col_letter} ({header}):")
        print(f"    Value: {cell.value}")
        print(f"    Data Type: {cell.data_type}")
        if cell.data_type == 'f':
            print(f"    Formula: {cell.value}")

print("\n" + "=" * 80)
print("پایان تحلیل")
print("=" * 80)

wb.close()
